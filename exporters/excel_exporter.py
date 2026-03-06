"""
Excel Exporter Module

This module implements Excel export functionality using openpyxl
with support for multiple worksheets, formatting, and charts.
"""

import asyncio
import os
from typing import Dict, List, Optional, Any, Union
from datetime import datetime, date
import logging

try:
    import openpyxl
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Border, Side, Alignment, Protection
    from openpyxl.utils import get_column_letter
    from openpyxl.chart import BarChart, Reference, Series
    EXCEL_AVAILABLE = True
except ImportError:
    EXCEL_AVAILABLE = False

from .base import BaseExporter, ExportConfig, ExportResult, ExportError


class ExcelExporter(BaseExporter):
    """Excel exporter with comprehensive formatting options"""
    
    def __init__(self, config: Optional[ExportConfig] = None):
        """Initialize Excel exporter with configuration"""
        super().__init__(config)
        
        if not EXCEL_AVAILABLE:
            raise ImportError("openpyxl is required for Excel export. Install with: pip install openpyxl")
        
        self.format = 'xlsx'
        self.include_charts = False
        self.auto_filter = True
        self.freeze_panes = True
        self.protected_worksheet = False
        
        # Styling options
        self.header_font = Font(bold=True, size=12)
        self.header_fill = PatternFill(start_color="DDDDDD", end_color="DDDDDD")
        self.header_border = Border(
            left=Side(style='thin', color='000000'),
            right=Side(style='thin', color='000000'),
            top=Side(style='thin', color='000000'),
            bottom=Side(style='thin', color='000000')
        )
        self.data_alignment = Alignment(horizontal='left', vertical='center')
        
        # Column formatting
        self.column_formats = {}
        self.column_widths = {}
    
    async def export(self, data: List[Dict[str, Any]], config: Optional[ExportConfig] = None) -> ExportResult:
        """Export data to Excel format"""
        export_config = config or self.config
        start_time = asyncio.get_event_loop().time()
        
        try:
            # Prepare data
            prepared_data = self.prepare_data(data)
            
            if not prepared_data:
                return ExportResult(
                    success=True,
                    exporter_name=self.name,
                    format=export_config.format,
                    record_count=0,
                    export_time=0.0
                )
            
            # Get file path
            file_path = export_config.get_full_path()
            
            # Create Excel workbook
            await self._create_excel_file(file_path, prepared_data, export_config)
            
            # Get file size
            file_size = self.get_file_size(file_path)
            
            # Calculate export time
            export_time = asyncio.get_event_loop().time() - start_time
            
            return ExportResult(
                success=True,
                file_path=file_path,
                record_count=len(prepared_data),
                file_size=file_size,
                export_time=export_time,
                exporter_name=self.name,
                format=export_config.format,
                metadata={
                    'worksheets': 1,
                    'include_charts': self.include_charts,
                    'auto_filter': self.auto_filter,
                    'freeze_panes': self.freeze_panes
                }
            )
            
        except Exception as e:
            export_time = asyncio.get_event_loop().time() - start_time
            return ExportResult(
                success=False,
                exporter_name=self.name,
                format=export_config.format,
                export_time=export_time,
                error_message=str(e)
            )
    
    async def _create_excel_file(self, file_path: str, data: List[Dict[str, Any]], config: ExportConfig):
        """Create Excel workbook with data"""
        try:
            # Create workbook
            wb = Workbook()
            ws = wb.active
            ws.title = "Data"
            
            # Get column names
            if data:
                column_names = list(data[0].keys())
            else:
                column_names = []
            
            # Write headers
            if config.include_headers and column_names:
                await self._write_headers(ws, column_names)
            
            # Write data
            await self._write_data_rows(ws, data, column_names)
            
            # Apply formatting
            await self._apply_formatting(ws, column_names)
            
            # Add features
            if self.auto_filter:
                await self._add_auto_filter(ws, len(column_names))
            
            if self.freeze_panes:
                await self._freeze_panes(ws)
            
            if self.include_charts and len(data) > 1:
                await self._add_charts(wb, ws, data, column_names)
            
            if self.protected_worksheet:
                await self._protect_worksheet(ws)
            
            # Save workbook
            wb.save(file_path)
            
        except Exception as e:
            raise ExportError(f"Failed to create Excel file {file_path}: {str(e)}")
    
    async def _write_headers(self, ws, column_names: List[str]):
        """Write header row with styling"""
        for col_idx, column_name in enumerate(column_names, 1):
            cell = ws.cell(row=1, column=col_idx, value=column_name)
            cell.font = self.header_font
            cell.fill = self.header_fill
            cell.border = self.header_border
            cell.alignment = Alignment(horizontal='center', vertical='center')
    
    async def _write_data_rows(self, ws, data: List[Dict[str, Any]], column_names: List[str]):
        """Write data rows"""
        for row_idx, record in enumerate(data, start=2):
            for col_idx, column_name in enumerate(column_names, 1):
                value = record.get(column_name)
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                
                # Apply column-specific formatting
                if column_name in self.column_formats:
                    cell.number_format = self.column_formats[column_name]
                
                # Apply alignment
                cell.alignment = self.data_alignment
    
    async def _apply_formatting(self, ws, column_names: List[str]):
        """Apply column formatting and widths"""
        for col_idx, column_name in enumerate(column_names, 1):
            column_letter = get_column_letter(col_idx)
            
            # Set column width
            if column_name in self.column_widths:
                ws.column_dimensions[column_letter] = self.column_widths[column_name]
            else:
                # Auto-size column based on content
                ws.column_dimensions[column_letter] = max(len(column_name), 15)
    
    async def _add_auto_filter(self, ws, column_count: int):
        """Add auto-filter to data"""
        if column_count > 0:
            last_column = get_column_letter(column_count)
            ws.auto_filter.ref = f"A1:{last_column}{len(ws['A'])}"
    
    async def _freeze_panes(self, ws):
        """Freeze header row"""
        ws.freeze_panes = "A2"
    
    async def _add_charts(self, wb, ws, data: List[Dict[str, Any]], column_names: List[str]):
        """Add charts to workbook"""
        try:
            # Find numeric columns for charting
            numeric_columns = []
            for col_idx, column_name in enumerate(column_names):
                if any(record.get(column_name) for record in data if record.get(column_name) is not None):
                    try:
                        # Check if column contains numeric data
                        sample_values = [record.get(column_name) for record in data[:10] if record.get(column_name) is not None]
                        if any(isinstance(v, (int, float)) for v in sample_values):
                            numeric_columns.append((col_idx, column_name))
                    except:
                        continue
            
            # Create bar chart for first numeric column
            if len(numeric_columns) > 0:
                col_idx, col_name = numeric_columns[0]
                await self._create_bar_chart(wb, ws, col_idx, col_name, data, column_names)
                
        except Exception as e:
            self.logger.warning(f"Failed to add charts: {str(e)}")
    
    async def _create_bar_chart(self, wb, ws, col_idx: int, column_name: str, data: List[Dict[str, Any]], column_names: List[str]):
        """Create a bar chart for a numeric column"""
        try:
            # Create chart object
            chart = BarChart()
            
            # Define chart data
            data_ref = Reference(ws, min_row=2, max_col=len(column_names), min_col=1)
            categories_ref = Reference(ws, min_row=2, max_col=len(column_names), min_col=1)
            
            # Add series (using the numeric column)
            values = []
            labels = []
            
            for record in data:
                value = record.get(column_name)
                if value is not None:
                    values.append(value)
                    labels.append(str(record.get(column_names[0], '')))  # Use first column as label
            
            if values:
                series = Series(values, title=column_name)
                chart.append(series)
            
            # Set chart properties
            chart.title = f"{column_name} Chart"
            chart.style = 10
            chart.x_axis.title = "Records"
            chart.y_axis.title = column_name
            
            # Add chart to worksheet
            ws.add_chart(chart)
            
            # Position chart
            chart.set_position('E2')
            
        except Exception as e:
            self.logger.warning(f"Failed to create bar chart: {str(e)}")
    
    async def _protect_worksheet(self, ws):
        """Protect worksheet from modifications"""
        ws.protection.sheet = True
        ws.protection.password = "password"  # You might want to make this configurable
    
    def set_column_format(self, column_name: str, format_code: str):
        """Set number format for a column"""
        self.column_formats[column_name] = format_code
    
    def set_column_width(self, column_name: str, width: float):
        """Set width for a column"""
        self.column_widths[column_name] = width
    
    def set_header_style(self, font: Optional[Font] = None, fill: Optional[PatternFill] = None, border: Optional[Border] = None):
        """Set header styling"""
        if font:
            self.header_font = font
        if fill:
            self.header_fill = fill
        if border:
            self.header_border = border
    
    def set_data_alignment(self, alignment: Alignment):
        """Set data alignment"""
        self.data_alignment = alignment
    
    def enable_charts(self, enable: bool = True):
        """Enable or disable chart creation"""
        self.include_charts = enable
    
    def enable_auto_filter(self, enable: bool = True):
        """Enable or disable auto-filter"""
        self.auto_filter = enable
    
    def enable_freeze_panes(self, enable: bool = True):
        """Enable or disable pane freezing"""
        self.freeze_panes = enable
    
    def enable_protection(self, enable: bool = True, password: str = "password"):
        """Enable or disable worksheet protection"""
        self.protected_worksheet = enable
        if enable:
            # You might want to store the password securely
            self.protection_password = password


class MultiSheetExcelExporter(ExcelExporter):
    """Excel exporter with support for multiple worksheets"""
    
    def __init__(self, config: Optional[ExportConfig] = None):
        """Initialize multi-sheet Excel exporter"""
        super().__init__(config)
        self.sheet_configs = {}
    
    def add_sheet_config(self, sheet_name: str, data_filter: callable = None, 
                         styling: Optional[Dict[str, Any]] = None):
        """Add configuration for a specific worksheet"""
        self.sheet_configs[sheet_name] = {
            'data_filter': data_filter,
            'styling': styling or {}
        }
    
    async def export(self, data: List[Dict[str, Any]], config: Optional[ExportConfig] = None) -> ExportResult:
        """Export data to Excel with multiple worksheets"""
        export_config = config or self.config
        start_time = asyncio.get_event_loop().time()
        
        try:
            # Prepare data
            prepared_data = self.prepare_data(data)
            
            if not prepared_data:
                return ExportResult(
                    success=True,
                    exporter_name=self.name,
                    format=export_config.format,
                    record_count=0,
                    export_time=0.0
                )
            
            # Get file path
            file_path = export_config.get_full_path()
            
            # Create Excel workbook with multiple sheets
            await self._create_multi_sheet_excel(file_path, prepared_data, export_config)
            
            # Get file size
            file_size = self.get_file_size(file_path)
            
            # Calculate export time
            export_time = asyncio.get_event_loop().time() - start_time
            
            return ExportResult(
                success=True,
                file_path=file_path,
                record_count=len(prepared_data),
                file_size=file_size,
                export_time=export_time,
                exporter_name=self.name,
                format=export_config.format,
                metadata={
                    'worksheets': len(self.sheet_configs) + 1,  # +1 for default sheet
                    'multi_sheet': True
                }
            )
            
        except Exception as e:
            export_time = asyncio.get_event_loop().time() - start_time
            return ExportResult(
                success=False,
                exporter_name=self.name,
                format=export_config.format,
                export_time=export_time,
                error_message=str(e)
            )
    
    async def _create_multi_sheet_excel(self, file_path: str, data: List[Dict[str, Any]], config: ExportConfig):
        """Create Excel workbook with multiple worksheets"""
        try:
            wb = Workbook()
            
            # Create default sheet with all data
            ws_default = wb.active
            ws_default.title = "All Data"
            await self._create_sheet_data(ws_default, data, config, self.sheet_configs.get('default', {}))
            
            # Create additional sheets
            for sheet_name, sheet_config in self.sheet_configs.items():
                ws = wb.create_sheet(title=sheet_name)
                
                # Apply data filter if specified
                if sheet_config['data_filter']:
                    filtered_data = [record for record in data if sheet_config['data_filter'](record)]
                else:
                    filtered_data = data
                
                await self._create_sheet_data(ws, filtered_data, config, sheet_config)
            
            # Save workbook
            wb.save(file_path)
            
        except Exception as e:
            raise ExportError(f"Failed to create multi-sheet Excel file {file_path}: {str(e)}")
    
    async def _create_sheet_data(self, ws, data: List[Dict[str, Any]], config: ExportConfig, sheet_config: Dict[str, Any]):
        """Create data for a specific worksheet"""
        # Apply sheet-specific styling
        styling = sheet_config.get('styling', {})
        
        # Override default styling if specified
        original_header_font = self.header_font
        original_header_fill = self.header_fill
        original_header_border = self.header_border
        
        try:
            if 'header_font' in styling:
                self.header_font = styling['header_font']
            if 'header_fill' in styling:
                self.header_fill = styling['header_fill']
            if 'header_border' in styling:
                self.header_border = styling['header_border']
            
            # Create sheet data
            if data:
                column_names = list(data[0].keys())
                
                # Write headers
                if config.include_headers:
                    await self._write_headers(ws, column_names)
                
                # Write data rows
                await self._write_data_rows(ws, data, column_names)
                
                # Apply formatting
                await self._apply_formatting(ws, column_names)
                
                # Add features
                if self.auto_filter:
                    await self._add_auto_filter(ws, len(column_names))
                
                if self.freeze_panes:
                    await self._freeze_panes(ws)
            
        finally:
            # Restore original styling
            self.header_font = original_header_font
            self.header_fill = original_header_fill
            self.header_border = original_header_border


# Convenience functions for common Excel export scenarios
def export_to_excel(data: List[Dict[str, Any]], filename: str, **kwargs) -> ExportResult:
    """Quick Excel export function"""
    config = ExportConfig(filename=filename, format='xlsx', **kwargs)
    exporter = ExcelExporter(config)
    return exporter.export_with_stats(data, config)


def create_excel_exporter(pretty_print: bool = True, auto_filter: bool = True, freeze_panes: bool = True) -> ExcelExporter:
    """Create Excel exporter with common settings"""
    config = ExportConfig(format='xlsx')
    exporter = ExcelExporter(config)
    exporter.enable_auto_filter(auto_filter)
    exporter.enable_freeze_panes(freeze_panes)
    return exporter


def create_multi_sheet_excel_exporter() -> MultiSheetExcelExporter:
    """Create multi-sheet Excel exporter"""
    return MultiSheetExcelExporter()


# Example usage patterns
def example_usage():
    """Example usage patterns for Excel export"""
    
    # Pattern 1: Basic Excel export
    data = [
        {'name': 'John', 'age': 30, 'salary': 50000.50, 'department': 'Engineering'},
        {'name': 'Jane', 'age': 25, 'salary': 60000.00, 'department': 'Marketing'},
        {'name': 'Bob', 'age': 35, 'salary': 75000.75, 'department': 'Engineering'}
    ]
    
    result = export_to_excel(data, 'employees.xlsx')
    
    # Pattern 2: Advanced Excel export with formatting
    exporter = create_excel_exporter()
    
    # Set column formats
    exporter.set_column_format('salary', '#,##0.00')
    exporter.set_column_format('age', '0')
    
    # Set column widths
    exporter.set_column_width('name', 20)
    exporter.set_column_width('salary', 15)
    
    # Enable charts
    exporter.enable_charts(True)
    
    result = exporter.export_with_stats(data)
    
    # Pattern 3: Multi-sheet export
    multi_exporter = create_multi_sheet_excel_exporter()
    
    # Add sheet for each department
    departments = set(record['department'] for record in data)
    for dept in departments:
        multi_exporter.add_sheet_config(
            sheet_name=dept,
            data_filter=lambda record: record.get('department') == dept
        )
    
    multi_result = multi_exporter.export_with_stats(data)
    
    print(f"Excel Export Result: {result.to_dict()}")
    print(f"Multi-Sheet Excel Export Result: {multi_result.to_dict()}")


if __name__ == "__main__":
    # Test Excel exporter
    logging.basicConfig(level=logging.INFO)
    
    # Create test data
    test_data = [
        {'name': 'John Doe', 'email': 'john@example.com', 'age': 30, 'salary': 50000.50, 'department': 'Engineering', 'join_date': datetime(2020, 1, 15)},
        {'name': 'Jane Smith', 'email': 'jane@example.com', 'age': 25, 'salary': 60000.00, 'department': 'Marketing', 'join_date': datetime(2019, 3, 20)},
        {'name': 'Bob Johnson', 'email': 'bob@example.com', 'age': 35, 'salary': 75000.75, 'department': 'Engineering', 'join_date': datetime(2018, 6, 10)}
    ]
    
    # Create and run Excel exporter
    exporter = create_excel_exporter()
    exporter.set_column_format('salary', '#,##0.00')
    exporter.set_column_format('age', '0')
    exporter.enable_charts(True)
    
    result = exporter.export_with_stats(test_data)
    
    print(f"Excel Export Result: {result.to_dict()}")
    
    # Test multi-sheet exporter
    multi_exporter = create_multi_sheet_excel_exporter()
    
    # Add sheet for each department
    departments = set(record['department'] for record in test_data)
    for dept in departments:
        multi_exporter.add_sheet_config(
            sheet_name=dept,
            data_filter=lambda record: record.get('department') == dept,
            styling={'header_fill': PatternFill(start_color="E6F3FF")}
        )
    
    multi_result = multi_exporter.export_with_stats(test_data)
    
    print(f"Multi-Sheet Excel Export Result: {multi_result.to_dict()}")
